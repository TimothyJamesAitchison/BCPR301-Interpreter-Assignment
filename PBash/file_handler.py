from __future__ import print_function
import sys
import csv
import re
import openpyxl
from validator import Validator


class FileHandler:
    def __init__(self, new_validator):
        self.validator = new_validator

    def open(self, file_path):
        if re.search(r'\.csv$', file_path):
            return self.csv_dict_reader(file_path)
        elif re.search(r'\.txt$', file_path):
            return self.txt_dict_reader(file_path)
        elif re.search(r'\.xlsx$', file_path):
            result = self.excel_reader(file_path)
            if result and self.validator.check_data_set(result):
                return result
            else:
                print("There were no valid entries in the file", file=sys.stderr)
                return False
        else:
            print('Invalid file extension', file=sys.stderr)
            return False

    def txt_dict_reader(self, filename):
        try:
            file = open(filename, "r")
        except FileNotFoundError:
            print('The file was not found', file=sys.stderr)
            return False
        the_list = []
        for line in file:
            dictionary = {}
            entries = line.split(";")
            for entry in entries:
                if len(entry.split("=")) == 2:
                    key = entry.split("=")[0]
                    value = entry.split("=")[1]
                    value = value.rstrip('\n')
                    dictionary[key] = value
                else:
                    print('The file was in an invalid format', file=sys.stderr)
                    return False
            if self.validator.check_line(dictionary):
                the_list.append(dictionary)
            else:
                print('Entry failed validation', file=sys.stderr)
        if self.validator.check_data_set(the_list):
            return the_list
        else:
            print("There were no valid entries in the file", file=sys.stderr)
            return False

    # hasitha
    def csv_dict_reader(self, filename):
        """
        >>> f = FileHandler(Validator())
        >>> result = f.csv_dict_reader("data.csv")
        >>> print(result[0]['EMPID'])
        A001
        >>> print(result[0]['GENDER'])
        F
        >>> print(result[0]['AGE'])
        21
        >>> print(result[0]['SALES'])
        001
        >>> print(result[0]['BMI'])
        Normal
        >>> print(result[0]['BIRTHDAY'])
        1-1-1996
        >>> print(result[0]['SALARY'])
        12
        """
        try:
            with open(filename) as f_obj:
                reader = csv.DictReader(f_obj, delimiter=',')
                the_list = []
                for line in reader:
                    employee = dict()
                    employee["EMPID"] = line["emp_id"]
                    employee["GENDER"] = line["gender"]
                    employee["AGE"] = line["age"]
                    employee["SALES"] = line["sales"]
                    employee["BMI"] = line["bmi"]
                    employee["SALARY"] = line["salary"]
                    employee["BIRTHDAY"] = line["birthday"]

                    if self.validator.check_line(employee):
                        the_list.append(employee)
                    else:
                        print('Entry failed validation', file=sys.stderr)
                if self.validator.check_data_set(the_list):
                    return the_list
                else:
                    print("There were no valid entries in the file", file=sys.stderr)
                    return False
        except FileNotFoundError:
            print('The file was not found', file=sys.stderr)
            return False

    # Tim
    @staticmethod
    def open_rules():
        try:
            file = open('rules.txt', "r")
        except FileNotFoundError:
            print('Cannot find rules.txt', file=sys.stderr)
            return False
        rules = {}
        for line in file:
            if len(line.split("=")) == 2:
                key = line.split("=")[0]
                value = line.split("=")[1]
                value = value.rstrip('\n')
                rules[key] = value
            else:
                print('The file was in an invalid format', file=sys.stderr)
                return False
        return rules

    # Tim
    def set_rules(self):
        self.validator.set_rules(self.open_rules())

    # Rosemary
    @staticmethod
    def open_help(help_command):
        """
        >>> f = FileHandler(new_validator=Validator)
        >>> print(f.open_help('line'))
        line command vitualize the data.
        >>> print(f.open_help('help'))
        help command brings out all command.
        >>> print(f.open_help('helpp'))
        No such command.
        """
        try:
            file = open("help.txt", "r")
            for line in file:
                if len(line.split("=")) == 2:
                    entries = line.split("=")
                    if help_command == entries[0]:
                        return entries[1].rstrip('\n')
                else:
                    print("Invalid help file format!")
        except FileNotFoundError:
            print('The help file was not found', file=sys.stderr)
        return "No such command."
        
    def excel_reader(self, filename):
        """
        >>> f = FileHandler(Validator())
        >>> result = f.excel_reader("testingdata.xlsx")
        >>> print(result[0]['EMPID'])
        A001
        >>> print(result[0]['GENDER'])
        F
        >>> print(result[0]['AGE'])
        21
        >>> print(result[0]['SALES'])
        001
        >>> print(result[0]['BMI'])
        Normal
        >>> print(result[0]['BIRTHDAY'])
        1-1-1996
        >>> print(result[0]['SALARY'])
        12
        """
        try:
            wb = openpyxl.load_workbook(filename)
            sheet = wb.active
            the_list = []
            for x in range(2, 29):
                employee = dict()
                employee["EMPID"] = sheet.cell(column=1, row=x).value
                employee["GENDER"] = sheet.cell(column=2, row=x).value
                employee["AGE"] = sheet.cell(column=3, row=x).value
                employee["SALES"] = sheet.cell(column=4, row=x).value
                employee["BMI"] = sheet.cell(column=5, row=x).value
                employee["SALARY"] = sheet.cell(column=6, row=x).value
                employee["BIRTHDAY"] = sheet.cell(column=7, row=x).value
                if self.validator.check_line(employee):
                    the_list.append(employee)
                else:
                    print('Entry failed validation', file=sys.stderr)
            if self.validator.check_data_set(the_list):
                return the_list
            else:
                print("There were no valid entries in the file", file=sys.stderr)
                return False
        except FileNotFoundError:
            print("File not found!", file=sys.stderr)
            return False

if __name__ == "__main__":
    import doctest
    doctest.testmod(verbose=1)
